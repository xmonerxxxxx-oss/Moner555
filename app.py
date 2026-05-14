import streamlit as st
import pandas as pd
import numpy as np
import io
import re
import google.generativeai as genai
from openpyxl.styles import PatternFill, Font
from groq import Groq

# ============================================================
# 1. إعداد الصفحة
# ============================================================
st.set_page_config(page_title="مساعد الصيدلية الذكي", layout="wide", page_icon="💊")

st.markdown("""
<style>
    .stChatMessage { border-radius: 12px; }
    .stDownloadButton button { background-color: #1f6feb; color: white; border-radius: 8px; }
    .block-container { padding-top: 2rem; }
    .stat-box { background:#f0f4ff; border-radius:10px; padding:10px 14px; margin:4px 0; border-left:4px solid #1f6feb; }
    .tip-box  { background:#fff8e1; border-radius:10px; padding:10px 14px; margin:4px 0; border-left:4px solid #f9a825; font-size:0.85rem; }
</style>
""", unsafe_allow_html=True)

# ============================================================
# 2. إعداد الـ API
# ============================================================
try:
    GEMINI_KEY = st.secrets["GEMINI_API_KEY"]
except Exception:
    st.error("🔑 لم يتم العثور على GEMINI_API_KEY في الـ Secrets.")
    st.stop()

try:
    GROQ_KEY = st.secrets["GROQ_API_KEY"]
except Exception:
    GROQ_KEY = None

# تتبع الـ provider الحالي
if "current_provider" not in st.session_state:
    st.session_state.current_provider = "gemini"

def call_ai(prompt: str) -> str:
    """استدعاء النموذج مع fallback تلقائي بين Gemini و Groq"""
    # حاول Gemini أولاً لو هو الحالي
    if st.session_state.current_provider == "gemini":
        try:
            genai.configure(api_key=GEMINI_KEY)
            model = genai.GenerativeModel("gemini-2.5-flash")
            result = model.generate_content(prompt).text
            return result
        except Exception as e:
            err = str(e).lower()
            if any(x in err for x in ["quota", "429", "rate", "limit", "exhausted"]):
                if GROQ_KEY:
                    st.toast("⚠️ Gemini وصل للحد — تحويل تلقائي لـ Groq...", icon="🔄")
                    st.session_state.current_provider = "groq"
                    # استكمل بـ Groq
                else:
                    raise Exception("quota_gemini")
            else:
                raise

    # Groq
    if st.session_state.current_provider == "groq":
        try:
            client = Groq(api_key=GROQ_KEY)
            response = client.chat.completions.create(
                model="llama-3.3-70b-versatile",
                messages=[{"role": "user", "content": prompt}],
                max_tokens=4096,
                temperature=0.1,
            )
            result = response.choices[0].message.content
            # جرب ترجع لـ Gemini في الطلب الجاي
            st.session_state.current_provider = "gemini"
            return result
        except Exception as e:
            err = str(e).lower()
            if any(x in err for x in ["quota", "429", "rate", "limit"]):
                raise Exception("quota_groq")
            raise

MAX_FILE_BYTES = 20 * 1024 * 1024

# ============================================================
# 3. تهيئة الـ Session State
# ============================================================
for key, default in [("history",[]),("dataframes",{}),("chat_messages",[]),("request_count",0)]:
    if key not in st.session_state:
        st.session_state[key] = default

# ============================================================
# 4. دوال مساعدة
# ============================================================
def extract_code(text: str) -> str:
    for pattern in [r"```python\s*(.*?)\s*```", r"```\s*(.*?)\s*```"]:
        match = re.search(pattern, text, re.DOTALL)
        if match:
            code = match.group(1).strip()
            return code[6:].strip() if code.startswith("python") else code
    cleaned = text.strip()
    return cleaned[6:].strip() if cleaned.startswith("python") else cleaned

def build_metadata(dataframes: dict) -> str:
    if not dataframes:
        return "لا توجد ملفات مرفوعة."
    parts = []
    for name, df in dataframes.items():
        col_info = [f"{c} ({str(df[c].dtype)}, {int(df[c].isna().sum())} فارغ)" for c in df.columns]
        info = (
            f"📄 الملف: {name}\n"
            f"   الصفوف: {len(df):,} | الأعمدة: {len(df.columns)}\n"
            f"   الأعمدة: {', '.join(df.columns.tolist())}\n"
            f"   تفاصيل:\n   " + "\n   ".join(col_info) + "\n"
            f"   عينة (أول 3 صفوف):\n{df.head(3).to_string(index=False)}\n"
        )
        parts.append(info)
    return "\n".join(parts)

def safe_exec(code: str, dataframes: dict):
    import openpyxl
    g = {
        "__builtins__": __builtins__,
        "pd": pd, "np": np, "io": io,
        "PatternFill": PatternFill, "Font": Font,
        "openpyxl": openpyxl,
        "dataframes": dataframes,
    }
    lc = {}
    try:
        exec(compile(code,"<ai>","exec"), g, lc)
    except Exception as e:
        return None, None, str(e)
    final_result = lc.get("final_result") if "final_result" in lc else g.get("final_result")
    apply_styling = lc.get("apply_styling") if "apply_styling" in lc else g.get("apply_styling")
    return final_result, apply_styling, None

def create_excel_bytes(df, apply_styling=None):
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter
    out = io.BytesIO()
    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Sheet1")
        ws = writer.sheets["Sheet1"]
        # تنسيق عرض الأعمدة تلقائياً
        for col_idx, col in enumerate(df.columns, 1):
            max_len = max(df[col].astype(str).map(len).max(), len(str(col))) + 4
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len, 40)
        if callable(apply_styling):
            try:
                apply_styling(writer)
                # إجبار حفظ التنسيق
                writer.book.save(out)
            except Exception as e:
                st.warning(f"⚠️ تعذّر التنسيق: {e}")
    return out.getvalue()

def classify_error(err_text: str) -> str:
    if err_text == "quota_gemini":
        return "⏳ Gemini وصل للحد اليومي ومفيش Groq متاح. جرب بكرة."
    if err_text == "quota_groq":
        return "⏳ Groq وصل للحد كمان. انتظر شوية وجرب تاني."
    err_lower = err_text.lower()
    if "quota" in err_lower or "429" in err_text:
        return "⏳ تجاوزت الحد — جاري المحاولة بنموذج بديل..."
    if "api_key" in err_lower or any(c in err_text for c in ["400","401","403"]):
        return "🔑 مشكلة في الـ API Key — تأكد إن المفتاح صح في الـ Secrets."
    if "deadline" in err_lower or "timeout" in err_lower or "503" in err_text:
        return "⌛ انتهت مهلة الاتصال. جرب تاني."
    return f"❌ خطأ غير متوقع: `{err_text}`"

# ============================================================
# 5. الـ Prompts
# ============================================================
RULES = """أنت خبير Python متخصص في تحليل بيانات الصيدليات. دقتك مهمة جداً لأن البيانات تخص صيدلية كبيرة.

قواعد صارمة يجب اتباعها بدقة تامة:

## الكود:
1. اكتب فقط كود Python داخل ```python ... ``` بدون أي شرح خارجه.
2. dataframes قاموس جاهز — لا تعيد تعريفه أبداً. مثال: df = dataframes['file.xlsx']
3. الكود يجب أن ينتهي دائماً بـ final_result من نوع DataFrame.
4. استخدم pd.to_numeric(errors='coerce') مع كل عمود رقمي بدون استثناء.
5. لا تستخدم print() أو open() أو pd.read_excel() أبداً.
6. لا تكتب import — pd و np و io و PatternFill و Font كلها جاهزة مباشرة.
7. لو الملف مش موجود استخدم: df = list(dataframes.values())[0]

## الأرقام والحسابات:
8. التقريب: إذا طُلب تقريب لـ X خانة عشرية استخدم round(x, X) وليس تحويل لـ int.
   مثال تقريب لخانة واحدة: df['عمود'] = df['عمود'].round(1)
   مثال تقريب لخانتين:    df['عمود'] = df['عمود'].round(2)
9. القسمة: استخدم دائماً .replace(0, np.nan) قبل القسمة لتفادي القسمة على صفر.
   مثال: df['الحافز'] = df['أوردرات'] / df['ساعات'].replace(0, np.nan)
10. القيم السالبة في "مطلوب شهر" و"مطلوب 40 يوم" تعني "لا يحتاج طلب" وليست خطأ.
11. "رصيد انور" و"رصيد حدوته" رصيدان لفرعين مختلفين — لا تخلطهما.

## التلوين (مهم جداً):
12. إذا طُلب تلوين: عرّف apply_styling(writer) واتبع هذا الشكل بالظبط:
def apply_styling(writer):
    from openpyxl.styles import PatternFill
    ws = writer.sheets['Sheet1']
    yellow = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    green  = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')
    red    = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    orange = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        val = row[COLUMN_INDEX].value  # COLUMN_INDEX = رقم العمود (يبدأ من 0)
        if val is not None:
            fill = green if CONDITION else red
            for cell in row:
                cell.fill = fill
13. لوّن كل خلايا الصف وليس خلية واحدة فقط — استخدم: for cell in row: cell.fill = fill
14. تأكد إن COLUMN_INDEX صح بناءً على ترتيب الأعمدة في final_result (يبدأ من 0)."""

def build_prompt(query, metadata, history):
    recent = "\n".join(history[-4:]) if history else "لا يوجد سياق سابق."
    return f"{RULES}\n\n## الملفات:\n{metadata}\n\n## السياق:\n{recent}\n\n## الطلب:\n{query}\n\nاكتب الكود:"

def build_fix_prompt(code, error, metadata):
    return f"{RULES}\n\nالكود:\n```python\n{code}\n```\nالخطأ: {error}\nالملفات:\n{metadata}\n\nأصلح الكود كاملاً داخل ```python ... ``` فقط."

# ============================================================
# 6. الواجهة
# ============================================================
st.title("💊 مساعد الصيدلية الذكي")
st.caption("ارفع ملفات Excel أو CSV وابدأ التحليل بالعربي")

with st.sidebar:
    st.header("📁 إدارة الملفات")
    uploaded_files = st.file_uploader("ارفع ملفاتك هنا", type=["xlsx","xls","csv"], accept_multiple_files=True)

    if uploaded_files:
        for file in uploaded_files:
            if file.size > MAX_FILE_BYTES:
                st.error(f"❌ {file.name} أكبر من 20 ميجا.")
                continue
            if file.name not in st.session_state.dataframes:
                try:
                    df = pd.read_csv(file, encoding="utf-8-sig") if file.name.endswith(".csv") else pd.read_excel(file)
                    st.session_state.dataframes[file.name] = df
                    st.success(f"✅ {file.name} ({len(df):,} صف)")
                except Exception as e:
                    st.error(f"❌ {file.name}: {e}")

    if st.session_state.dataframes:
        st.divider()
        st.subheader("📊 الملفات المحملة")
        for name, df in list(st.session_state.dataframes.items()):
            with st.expander(f"📄 {name}"):
                st.write(f"**الصفوف:** {len(df):,} | **الأعمدة:** {len(df.columns)}")
                st.write(f"**الأعمدة:** {', '.join(df.columns.tolist())}")
                if st.button("🗑️ حذف", key=f"del_{name}"):
                    del st.session_state.dataframes[name]
                    st.rerun()

    st.divider()
    provider_icon = '🟢 Gemini' if st.session_state.get('current_provider','gemini') == 'gemini' else '🟠 Groq'
    st.markdown(f'<div class="stat-box">🤖 النموذج الحالي: <b>{provider_icon}</b></div>', unsafe_allow_html=True)
    st.markdown(f'<div class="stat-box">📊 طلبات الجلسة: <b>{st.session_state.request_count}</b></div>', unsafe_allow_html=True)

    c1, c2 = st.columns(2)
    with c1:
        if st.button("🗑️ مسح الكل", use_container_width=True):
            st.session_state.dataframes = {}; st.session_state.chat_messages = []; st.session_state.history = []
            st.rerun()
    with c2:
        if st.button("💬 مسح المحادثة", use_container_width=True):
            st.session_state.chat_messages = []; st.session_state.history = []
            st.rerun()

    st.divider()
    st.subheader("💡 أمثلة جاهزة")
    examples = [
        "وريني الأصناف اللي مطلوب شهر أكبر من صفر",
        "وريني الأصناف اللي رصيد انور صفر ومطلوب شهر موجب",
        "احسب إجمالي قيمة المطلوب شهر × الجمهور",
        "وريني أكثر 20 صنف مطلوباً هذا الشهر",
        "لون الصفوف اللي مطلوب شهر أكبر من 5 باللون الأخضر",
        "فلتر أصناف قسم الثلاجة فقط",
        "أزل الأصناف اللي رصيد انور ورصيد حدوته كلاهما صفر",
        "اعرض إحصائيات وصفية للبيانات",
    ]
    for ex in examples:
        if st.button(f"▶ {ex}", key=ex, use_container_width=True):
            st.session_state["preset_query"] = ex
            st.rerun()

    st.markdown('<div class="tip-box">💡 <b>تلميح:</b> اسأل بالعربي العادي زي ما بتتكلم</div>', unsafe_allow_html=True)

# ============================================================
# 7. عرض المحادثة
# ============================================================
for msg in st.session_state.chat_messages:
    with st.chat_message(msg["role"]):
        st.markdown(msg["content"])
        if msg.get("data") is not None:
            st.dataframe(msg["data"], use_container_width=True)
        if msg.get("code"):
            with st.expander("👨‍💻 الكود المُنفَّذ"):
                st.code(msg["code"], language="python")
        if msg.get("download_bytes") is not None:
            st.download_button("📥 تحميل النتيجة (Excel)", data=msg["download_bytes"],
                file_name="النتيجة.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key=f"dl_{msg.get('msg_id', id(msg))}")

# ============================================================
# 8. معالجة الإدخال
# ============================================================
preset = st.session_state.pop("preset_query", None)
user_query = st.chat_input("اكتب طلبك هنا... مثال: وريني الأصناف اللي رصيدها صفر") or preset

if user_query:
    with st.chat_message("user"):
        st.markdown(user_query)
    st.session_state.chat_messages.append({"role":"user","content":user_query})
    st.session_state.history.append(f"المستخدم: {user_query}")

    if not st.session_state.dataframes:
        msg = "⚠️ لم يتم رفع أي ملفات. يرجى رفع ملف من الشريط الجانبي."
        with st.chat_message("assistant"): st.markdown(msg)
        st.session_state.chat_messages.append({"role":"assistant","content":msg})
        st.stop()

    metadata = build_metadata(st.session_state.dataframes)

    with st.chat_message("assistant"):
        status = st.empty()
        status.markdown("🧠 جاري التحليل وكتابة الكود...")
        msg_data = {"role":"assistant","content":"","msg_id":len(st.session_state.chat_messages)}

        try:
            raw = call_ai(build_prompt(user_query, metadata, st.session_state.history))
            code = extract_code(raw)
            st.session_state.request_count += 1

            if not code.strip():
                raise ValueError("لم يتم استخراج كود من رد الذكاء الاصطناعي.")

            final_result, apply_styling, exec_error = safe_exec(code, st.session_state.dataframes)

            if exec_error:
                status.markdown(f"⚠️ جاري الإصلاح التلقائي... `{exec_error}`")
                fixed_raw = call_ai(build_fix_prompt(code, exec_error, metadata))
                fixed_code = extract_code(fixed_raw)
                st.session_state.request_count += 1
                final_result, apply_styling, exec_error2 = safe_exec(fixed_code, st.session_state.dataframes)

                if exec_error2:
                    err_msg = f"❌ فشل التنفيذ حتى بعد الإصلاح.\n\n**الخطأ:** `{exec_error2}`"
                    status.markdown(err_msg)
                    with st.expander("👨‍💻 الكود الأخير"): st.code(fixed_code, language="python")
                    msg_data["content"] = err_msg; msg_data["code"] = fixed_code
                    st.session_state.chat_messages.append(msg_data)
                    st.stop()

                code = fixed_code
                st.toast("✅ تم إصلاح الكود تلقائياً!")

            if isinstance(final_result, pd.DataFrame):
                rows, cols = len(final_result), len(final_result.columns)
                success_msg = f"✅ تمت المعالجة! **{rows:,} صف × {cols} عمود**"
                status.markdown(success_msg)
                st.dataframe(final_result, use_container_width=True)
                excel_bytes = create_excel_bytes(final_result, apply_styling)
                st.download_button("📥 تحميل النتيجة (Excel)", data=excel_bytes,
                    file_name="النتيجة.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key=f"dl_new_{len(st.session_state.chat_messages)}")
                with st.expander("👨‍💻 الكود المُنفَّذ"): st.code(code, language="python")
                msg_data.update({"content":success_msg,"data":final_result,"download_bytes":excel_bytes,"code":code})
                st.session_state.history.append(f"المساعد: تمت المعالجة — {rows} صف، أعمدة: {final_result.columns.tolist()}")

            elif final_result is not None:
                warn = f"⚠️ final_result ليس DataFrame (النوع: {type(final_result).__name__})."
                status.markdown(warn); st.write(final_result)
                with st.expander("👨‍💻 الكود"): st.code(code, language="python")
                msg_data["content"] = warn; msg_data["code"] = code
            else:
                warn = "⚠️ الكود نُفِّذ لكن لم يتم تعريف final_result."
                status.markdown(warn)
                with st.expander("👨‍💻 الكود"): st.code(code, language="python")
                msg_data["content"] = warn; msg_data["code"] = code

        except Exception as e:
            err_msg = classify_error(str(e))
            status.markdown(err_msg)
            msg_data["content"] = err_msg

        st.session_state.chat_messages.append(msg_data)

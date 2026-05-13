import streamlit as st
import pandas as pd
import numpy as np
import io
import re
import google.generativeai as genai
from openpyxl.styles import PatternFill, Font

# ============================================================
# 1. إعداد الصفحة
# ============================================================
st.set_page_config(page_title="مساعد الصيدلية الذكي", layout="wide", page_icon="💊")

st.markdown("""
<style>
    .stChatMessage { border-radius: 12px; }
    .stDownloadButton button { background-color: #1f6feb; color: white; border-radius: 8px; }
    .block-container { padding-top: 2rem; }
    .stat-box { background:#f0f4ff; border-radius:10px; padding:10px 14px; margin:4px 0; border-left:4px solid #1f6feb; }
    .tip-box  { background:#fff8e1; border-radius:10px; padding:10px 14px; margin:4px 0; border-left:4px solid #f9a825; font-size:0.85rem; }
</style>
""", unsafe_allow_html=True)

# ============================================================
# 2. إعداد الـ API
# ============================================================
API_KEY = st.secrets["GEMINI_API_KEY"]

def get_model():
    genai.configure(api_key=API_KEY)
    return genai.GenerativeModel("gemini-2.5-flash")

MAX_FILE_BYTES = 20 * 1024 * 1024

# ============================================================
# 3. تهيئة الـ Session State
# ============================================================
for key, default in [("history",[]),("dataframes",{}),("chat_messages",[]),("request_count",0)]:
    if key not in st.session_state:
        st.session_state[key] = default

# ============================================================
# 4. دوال مساعدة
# ============================================================
def extract_code(text: str) -> str:
    for pattern in [r"```python\s*(.*?)\s*```", r"```\s*(.*?)\s*```"]:
        match = re.search(pattern, text, re.DOTALL)
        if match:
            code = match.group(1).strip()
            return code[6:].strip() if code.startswith("python") else code
    cleaned = text.strip()
    return cleaned[6:].strip() if cleaned.startswith("python") else cleaned

def build_metadata(dataframes: dict) -> str:
    if not dataframes:
        return "لا توجد ملفات مرفوعة."
    parts = []
    for name, df in dataframes.items():
        col_info = [f"{c} ({str(df[c].dtype)}, {int(df[c].isna().sum())} فارغ)" for c in df.columns]
        info = (
            f"📄 الملف: {name}\n"
            f"   الصفوف: {len(df):,} | الأعمدة: {len(df.columns)}\n"
            f"   الأعمدة: {', '.join(df.columns.tolist())}\n"
            f"   تفاصيل:\n   " + "\n   ".join(col_info) + "\n"
            f"   عينة (أول 3 صفوف):\n{df.head(3).to_string(index=False)}\n"
        )
        parts.append(info)
    return "\n".join(parts)

def safe_exec(code: str, dataframes: dict):
    allowed = {
        "len":len,"range":range,"enumerate":enumerate,"zip":zip,"list":list,"dict":dict,
        "set":set,"tuple":tuple,"str":str,"int":int,"float":float,"bool":bool,
        "print":lambda*a,**k:None,"isinstance":isinstance,"type":type,
        "min":min,"max":max,"sum":sum,"abs":abs,"round":round,"sorted":sorted,
        "reversed":reversed,"map":map,"filter":filter,"any":any,"all":all,
        "hasattr":hasattr,"getattr":getattr,
    }
    g = {"__builtins__":allowed,"pd":pd,"np":np,"io":io,"PatternFill":PatternFill,"Font":Font,"dataframes":dataframes}
    lc = {}
    try:
        exec(compile(code,"<ai>","exec"), g, lc)
    except Exception as e:
        return None, None, str(e)
    return (lc.get("final_result") or g.get("final_result")), (lc.get("apply_styling") or g.get("apply_styling")), None

def create_excel_bytes(df, apply_styling=None):
    out = io.BytesIO()
    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Sheet1")
        if callable(apply_styling):
            try: apply_styling(writer)
            except Exception as e: st.warning(f"⚠️ تعذّر التنسيق: {e}")
    return out.getvalue()

# ============================================================
# 5. الـ Prompts
# ============================================================
RULES = """أنت خبير Python متخصص في تحليل بيانات الصيدليات باستخدام pandas و numpy.

قواعد صارمة:
1. اكتب فقط كود Python داخل ```python ... ``` بدون أي شرح خارجه.
2. dataframes قاموس جاهز — لا تعيد تعريفه. مثال: df = dataframes['file.xlsx']
3. الكود يجب أن ينتهي بـ final_result من نوع DataFrame.
4. استخدم pd.to_numeric(errors='coerce') دايماً مع الأرقام.
5. القيم السالبة في "مطلوب شهر" و"مطلوب 40 يوم" تعني "لا يحتاج طلب" وليست خطأ.
6. "رصيد انور" و"رصيد حدوته" رصيدان لفرعين مختلفين.
7. إذا طُلب تنسيق: عرّف apply_styling(writer) — الشيت دايماً 'Sheet1'.
   yellow=PatternFill(start_color='FFFF00',end_color='FFFF00',fill_type='solid')
   ws=writer.sheets['Sheet1']
8. لا تستخدم print() أو open() أو pd.read_excel()."""

def build_prompt(query, metadata, history):
    recent = "\n".join(history[-4:]) if history else "لا يوجد سياق سابق."
    return f"{RULES}\n\n## الملفات:\n{metadata}\n\n## السياق:\n{recent}\n\n## الطلب:\n{query}\n\nاكتب الكود:"

def build_fix_prompt(code, error, metadata):
    return f"{RULES}\n\nالكود:\n```python\n{code}\n```\nالخطأ: {error}\nالملفات:\n{metadata}\n\nأصلح الكود كاملاً داخل ```python ... ``` فقط."

# ============================================================
# 6. الواجهة
# ============================================================
st.title("💊 مساعد الصيدلية الذكي")
st.caption("ارفع ملفات Excel أو CSV وابدأ التحليل بالعربي")

with st.sidebar:
    st.header("📁 إدارة الملفات")
    uploaded_files = st.file_uploader("ارفع ملفاتك هنا", type=["xlsx","xls","csv"], accept_multiple_files=True)

    if uploaded_files:
        for file in uploaded_files:
            if file.size > MAX_FILE_BYTES:
                st.error(f"❌ {file.name} أكبر من 20 ميجا.")
                continue
            if file.name not in st.session_state.dataframes:
                try:
                    df = pd.read_csv(file, encoding="utf-8-sig") if file.name.endswith(".csv") else pd.read_excel(file)
                    st.session_state.dataframes[file.name] = df
                    st.success(f"✅ {file.name} ({len(df):,} صف)")
                except Exception as e:
                    st.error(f"❌ {file.name}: {e}")

    if st.session_state.dataframes:
        st.divider()
        st.subheader("📊 الملفات المحملة")
        for name, df in list(st.session_state.dataframes.items()):
            with st.expander(f"📄 {name}"):
                st.write(f"**الصفوف:** {len(df):,} | **الأعمدة:** {len(df.columns)}")
                st.write(f"**الأعمدة:** {', '.join(df.columns.tolist())}")
                if st.button("🗑️ حذف", key=f"del_{name}"):
                    del st.session_state.dataframes[name]
                    st.rerun()

    st.divider()
    st.markdown(f'<div class="stat-box">📊 طلبات الجلسة: <b>{st.session_state.request_count}</b></div>', unsafe_allow_html=True)

    c1, c2 = st.columns(2)
    with c1:
        if st.button("🗑️ مسح الكل", use_container_width=True):
            st.session_state.dataframes = {}; st.session_state.chat_messages = []; st.session_state.history = []
            st.rerun()
    with c2:
        if st.button("💬 مسح المحادثة", use_container_width=True):
            st.session_state.chat_messages = []; st.session_state.history = []
            st.rerun()

    st.divider()
    st.subheader("💡 أمثلة جاهزة")
    examples = [
        "وريني الأصناف اللي مطلوب شهر أكبر من صفر",
        "وريني الأصناف اللي رصيد انور صفر ومطلوب شهر موجب",
        "احسب إجمالي قيمة المطلوب شهر × الجمهور",
        "وريني أكثر 20 صنف مطلوباً هذا الشهر",
        "لون الصفوف اللي مطلوب شهر أكبر من 5 باللون الأخضر",
        "فلتر أصناف قسم الثلاجة فقط",
        "أزل الأصناف اللي رصيد انور ورصيد حدوته كلاهما صفر",
        "اعرض إحصائيات وصفية للبيانات",
    ]
    for ex in examples:
        if st.button(f"▶ {ex}", key=ex, use_container_width=True):
            st.session_state["preset_query"] = ex
            st.rerun()

    st.markdown('<div class="tip-box">💡 <b>تلميح:</b> اسأل بالعربي العادي زي ما بتتكلم</div>', unsafe_allow_html=True)

# ============================================================
# 7. عرض المحادثة
# ============================================================
for msg in st.session_state.chat_messages:
    with st.chat_message(msg["role"]):
        st.markdown(msg["content"])
        if msg.get("data") is not None:
            st.dataframe(msg["data"], use_container_width=True)
        if msg.get("code"):
            with st.expander("👨‍💻 الكود المُنفَّذ"):
                st.code(msg["code"], language="python")
        if msg.get("download_bytes") is not None:
            st.download_button("📥 تحميل النتيجة (Excel)", data=msg["download_bytes"],
                file_name="النتيجة.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key=f"dl_{msg.get('msg_id', id(msg))}")

# ============================================================
# 8. معالجة الإدخال
# ============================================================
preset = st.session_state.pop("preset_query", None)
user_query = st.chat_input("اكتب طلبك هنا... مثال: وريني الأصناف اللي رصيدها صفر") or preset

if user_query:
    with st.chat_message("user"):
        st.markdown(user_query)
    st.session_state.chat_messages.append({"role":"user","content":user_query})
    st.session_state.history.append(f"المستخدم: {user_query}")

    if not st.session_state.dataframes:
        msg = "⚠️ لم يتم رفع أي ملفات. يرجى رفع ملف من الشريط الجانبي."
        with st.chat_message("assistant"): st.markdown(msg)
        st.session_state.chat_messages.append({"role":"assistant","content":msg})
        st.rerun()

    metadata = build_metadata(st.session_state.dataframes)

    with st.chat_message("assistant"):
        status = st.empty()
        status.markdown("🧠 جاري التحليل وكتابة الكود...")
        msg_data = {"role":"assistant","content":"","msg_id":len(st.session_state.chat_messages)}

        try:
            model = get_model()

            # الاستدعاء الأول
            raw = model.generate_content(build_prompt(user_query, metadata, st.session_state.history)).text
            code = extract_code(raw)
            st.session_state.request_count += 1

            if not code:
                raise ValueError("لم يتم استخراج كود من رد الذكاء الاصطناعي.")

            final_result, apply_styling, exec_error = safe_exec(code, st.session_state.dataframes)

            # إصلاح تلقائي
            if exec_error:
                status.markdown(f"⚠️ جاري الإصلاح التلقائي... `{exec_error}`")
                fixed_raw = model.generate_content(build_fix_prompt(code, exec_error, metadata)).text
                fixed_code = extract_code(fixed_raw)
                st.session_state.request_count += 1
                final_result, apply_styling, exec_error2 = safe_exec(fixed_code, st.session_state.dataframes)

                if exec_error2:
                    err_msg = f"❌ فشل التنفيذ حتى بعد الإصلاح.\n\n**الخطأ:** `{exec_error2}`"
                    status.markdown(err_msg)
                    with st.expander("👨‍💻 الكود الأخير"): st.code(fixed_code, language="python")
                    msg_data["content"] = err_msg; msg_data["code"] = fixed_code
                    st.session_state.chat_messages.append(msg_data)
                    st.rerun()

                code = fixed_code
                st.toast("✅ تم إصلاح الكود تلقائياً!")

            # عرض النتيجة
            if isinstance(final_result, pd.DataFrame):
                rows, cols = len(final_result), len(final_result.columns)
                success_msg = f"✅ تمت المعالجة! **{rows:,} صف × {cols} عمود**"
                status.markdown(success_msg)
                st.dataframe(final_result, use_container_width=True)
                excel_bytes = create_excel_bytes(final_result, apply_styling)
                st.download_button("📥 تحميل النتيجة (Excel)", data=excel_bytes,
                    file_name="النتيجة.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key=f"dl_new_{len(st.session_state.chat_messages)}")
                with st.expander("👨‍💻 الكود المُنفَّذ"): st.code(code, language="python")
                msg_data.update({"content":success_msg,"data":final_result,"download_bytes":excel_bytes,"code":code})
                st.session_state.history.append(f"المساعد: تمت المعالجة — {rows} صف، أعمدة: {final_result.columns.tolist()}")

            elif final_result is not None:
                warn = f"⚠️ final_result ليس DataFrame (النوع: {type(final_result).__name__})."
                status.markdown(warn); st.write(final_result)
                with st.expander("👨‍💻 الكود"): st.code(code, language="python")
                msg_data["content"] = warn; msg_data["code"] = code
            else:
                warn = "⚠️ الكود نُفِّذ لكن لم يتم تعريف final_result."
                status.markdown(warn)
                with st.expander("👨‍💻 الكود"): st.code(code, language="python")
                msg_data["content"] = warn; msg_data["code"] = code

        except Exception as e:
            err_text = str(e)
            if "quota" in err_text.lower() or "429" in err_text:
                err_msg = "⏳ تجاوزت الحد اليومي المجاني لـ Gemini. جرب بكرة أو انتظر شوية."
            elif "api_key" in err_text.lower() or "400" in err_text:
                err_msg = "🔑 مشكلة في الـ API Key."
            else:
                err_msg = f"❌ خطأ: `{err_text}`"
            status.markdown(err_msg)
            msg_data["content"] = err_msg

        st.session_state.chat_messages.append(msg_data)
